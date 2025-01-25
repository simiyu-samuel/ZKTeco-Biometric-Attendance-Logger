import datetime
from openpyxl import load_workbook, Workbook
import http.client
import json
from zk import ZK, const
import time

# Constants
EXCEL_FILE = "attendance_log.xlsx"
VALID_ACTIONS = [0, 1]  # 0: Sign-in, 1: Sign-out

# ZKTeco Device Configuration
ZK_IP = "192.168.0.201"  # Replace with your ZKTeco device IP address
ZK_PORT = 4370           # Default port for ZKTeco devices

# Infobip SMS API Configuration
INFOBIP_API_URL = "/sms/2/text/advanced"
INFOBIP_AUTH = "e15cbf620046a7ab6d5f8ea1b7d38149-3ded325d-0355-4e55-8dd7-76721d598534"
INFOBIP_SENDER_ID = "447491163443"

# Function to fetch user details from ZKTeco device
def get_user_details(user_id):
    conn = None
    try:
        zk = ZK(ZK_IP, port=ZK_PORT, timeout=5)
        conn = zk.connect()
        conn.disable_device()  # Temporarily disable the device during communication

        # Fetch all users
        users = conn.get_users()
        for user in users:
            if user.user_id == str(user_id):  # Match user_id
                conn.enable_device()  # Re-enable the device
                return {"name": user.name, "phone": user.privilege}  # Replace privilege with phone if stored
        conn.enable_device()
        return {"name": "Unknown User", "phone": None}
    except Exception as e:
        print(f"Error communicating with ZKTeco device: {e}")
        return {"name": "Unknown User", "phone": None}
    finally:
        if conn:
            conn.disconnect()

# Function to fetch user phone numbers
def get_user_phone(user_id):
    user_details = get_user_details(user_id)
    return user_details.get("phone")

# Function to fetch user names
def get_user_name(user_id):
    user_details = get_user_details(user_id)
    return user_details.get("name")

# Function to send SMS using Infobip
def send_sms_infobip(message):
    try:
        conn = http.client.HTTPSConnection("api.infobip.com")
        payload = json.dumps({
            "messages": [
                {
                    "destinations": [{"to": "254714894934"}],
                    "from": INFOBIP_SENDER_ID,
                    "text": message
                }
            ]
        })
        headers = {
            'Authorization': f'App {INFOBIP_AUTH}',
            'Content-Type': 'application/json',
            'Accept': 'application/json'
        }
        conn.request("POST", INFOBIP_API_URL, payload, headers)
        res = conn.getresponse()
        data = res.read()
        print(f"Infobip SMS Response: {data.decode('utf-8')}")
    except Exception as e:
        print(f"Error sending SMS via Infobip: {e}")

# Function to log data into Excel
def log_to_excel(user_id, action):
    try:
        # Attempt to load the workbook
        try:
            workbook = load_workbook(EXCEL_FILE)
            sheet = workbook.active
        except FileNotFoundError:
            print(f"File '{EXCEL_FILE}' not found. Creating a new file.")
            workbook = Workbook()
            sheet = workbook.active
            # Add headers to the new file
            if action == 0:
                action = "Sign-in"
            else:
                action = "Sign-out"
            print(action)
            sheet.append(["User ID", "Name", "Act", "Timestamp"])

        # Check if the user ID already exists in the sheet
        user_exists = False
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == user_id and row[2] == action:  # Check for same user ID and action
                user_exists = True
                break

        if user_exists:
            print(f"User ID {user_id} with action {action} is already logged.")
            return  # Skip logging and sending SMS if already processed

        # Log the user action into the sheet
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        user_name = get_user_name(user_id)
        sheet.append([user_id, user_name, action, timestamp])
        workbook.save(EXCEL_FILE)
        print(f"Logged: User ID={user_id}, Name={user_name}, Action={action}, Timestamp={timestamp}")

        # Send SMS notification
        phone_number = get_user_phone(user_id)
        if phone_number:
            message = f"Hello {user_name}, your {action} was recorded at {timestamp}."
            send_sms_infobip(message)
    except PermissionError:
        print(f"Permission denied. Please close the file '{EXCEL_FILE}' and try again.")
    except Exception as e:
        print(f"Error saving data to the file: {e}")

# Function to handle biometric events when a user is detected
def on_biometric_event(conn):
    print("Waiting for biometric event...")
    while True:
        # Check for any fingerprint match
        try:
            attendance = conn.get_attendance()
            if attendance:
                for record in attendance:
                    user_id = record.user_id
                    action = record.status  # Action can be 0 for sign-in, 1 for sign-out
                    if action in VALID_ACTIONS:
                        log_to_excel(user_id, action)
            time.sleep(1)  # Poll every second
        except Exception as e:
            print(f"Error reading attendance data: {e}")
            time.sleep(1)

# Main function to connect to the device and start monitoring
def main():
    try:
        print("Connecting to ZKTeco device...")
        zk = ZK(ZK_IP, port=ZK_PORT, timeout=5)
        conn = zk.connect()
        conn.disable_device()  # Temporarily disable the device for safe communication

        print("Connected successfully!")

        # Start monitoring the device for biometric events (fingerprint scans)
        on_biometric_event(conn)

    except Exception as e:
        print(f"Error: {e}")
    finally:
        if conn:
            conn.disconnect()
        print("Disconnected from ZKTeco device.")

if _name_ == "_main_":
    main()
