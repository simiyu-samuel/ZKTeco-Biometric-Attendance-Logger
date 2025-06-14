import datetime
import openpyxl
from openpyxl import load_workbook, Workbook
import http.client
import json
import ssl
from zk import ZK, const 
import time
import mysql.connector

# Constants
EXCEL_FILE = "attendance_log.xlsx"
VALID_ACTIONS = [0, 1]  # 0: Sign-in, 1: Sign-out

# ZKTeco Device Configuration
ZK_IP = "192.168.0.201"  # Replace with your ZKTeco device IP address
ZK_PORT = 4370  # Default port for ZKTeco devices

MSPACE_USERNAME = ""  # **REPLACE WITH YOUR mSpace USERNAME**
MSPACE_SENDER_ID = ""  # **REPLACE WITH YOUR mSpace SENDER ID**
MSPACE_API_KEY = ""  # **REPLACE WITH YOUR mSpace API KEY**

# Global variables for tracking recently logged users
recently_logged = set()
last_reset_date = None

# Database Configuration
db_config = {
    'host': 'localhost',
    'user': 'root',
    'password': '',
    'database': 'attendance_system'
}

# Function to reset the recently_logged set daily
def reset_recently_logged():
    global recently_logged, last_reset_date
    current_date = datetime.datetime.now().date()
    if last_reset_date != current_date:
        recently_logged.clear()
        last_reset_date = current_date
        print("Reset recently logged actions for the new day.")

# Function to fetch user details from ZKTeco device
def get_user_details(user_id):
    conn = None
    try:
        zk = ZK(ZK_IP, port=ZK_PORT, timeout=5)
        conn = zk.connect()
        conn.disable_device()  # Temporarily disable the device during communication

        users = conn.get_users()
        for user in users:
            if user.user_id == str(user_id):
                conn.enable_device()
                return {"name": user.name, "phone": user.privilege}  # Replace privilege with phone if stored
        conn.enable_device()
        return {"name": "Unknown User", "phone": None}
    except Exception as e:
        print(f"Error communicating with ZKTeco device: {e}")
        return {"name": "Unknown User", "phone": None}
    finally:
        if conn:
            conn.disconnect()

# Function to fetch user phone numbers
def get_user_phone(user_id):
    user_details = get_user_details(user_id)
    return user_details.get("phone")

# Function to fetch user names
def get_user_name(user_id):
    user_details = get_user_details(user_id)
    return user_details.get("name")

# Function to send SMS
def send_sms(message):
    context = ssl.create_default_context()
    context.check_hostname = False
    context.verify_mode = ssl.CERT_NONE
    
    conn = http.client.HTTPSConnection("api.mspace.co.ke", context=context)
    payload = json.dumps({
        "username": "Willie",
        "senderId": "ICORESYSTEM",
        "recipient": "254707735717",  # Adjust recipient as necessary
        "message": message
    })
    headers = {
        'apikey': MSPACE_API_KEY,
        'Content-Type': 'application/json'
    }
    
    try:
        conn.request("POST", "/smsapi/v2/sendtext", payload, headers)
        res = conn.getresponse()
        data = res.read()
        store_sms(message)  # Store the sent SMS log
        print(data.decode("utf-8"))
    except Exception as e:
        print(f"Error sending SMS: {e}")
    finally:
        conn.close()

# Function to store SMS logs in the database
def store_sms(message):
    connection = connect_to_db()
    if connection:
        cursor = connection.cursor()
        try:
            timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            query = "INSERT INTO sms_logs (message, timestamp) VALUES (%s, %s)"
            cursor.execute(query, (message, timestamp))
            connection.commit()
            print(f"Logged SMS: {message}, Timestamp={timestamp}")
        except mysql.connector.Error as err:
            print(f"Error saving SMS log to DB: {err}")
        finally:
            connection.close()

# Function to connect to the database
def connect_to_db():
    try:
        connection = mysql.connector.connect(**db_config)
        print('DB connection established.')
        return connection
    except mysql.connector.Error as err:
        print(f"Error occurred: {err}")
        return None

# Function to log data into the database
def log_to_db(user_id, user_name, action_str):
    connection = connect_to_db()
    if connection:
        cursor = connection.cursor()
        try:
            timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            query = "INSERT INTO attendance_logs (user_id, user_name, action, timestamp) VALUES (%s, %s, %s, %s)"
            cursor.execute(query, (user_id, user_name, action_str, timestamp))
            connection.commit()
            print(f"Logged to DB: User ID={user_id}, Name={user_name}, Action={action_str}, Timestamp={timestamp}")
        except mysql.connector.Error as err:
            print(f"Error saving attendance to DB: {err}")
        finally:
            connection.close()

# Function to log data into Excel
def log_to_excel_and_send_sms(user_id, action):
    global recently_logged
    try:
        reset_recently_logged()

        action_str = "Sign-in" if action == 0 else "Sign-out"
        current_date = datetime.datetime.now().date()
        unique_action = (user_id, action, current_date)

        if unique_action in recently_logged:
            print(f"Duplicate detected: User ID={user_id}, Action={action_str}, Date={current_date}")
            return

        recently_logged.add(unique_action)

        try:
            workbook = load_workbook(EXCEL_FILE)
            sheet = workbook.active
        except FileNotFoundError:
            print(f"File '{EXCEL_FILE}' not found. Creating a new file.")
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["User ID", "Name", "Action", "Timestamp"])

        timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        user_name = get_user_name(user_id)
        sheet.append([user_id, user_name, action_str, timestamp])
        workbook.save(EXCEL_FILE)
        print(f"Logged to Excel: User ID={user_id}, Name={user_name}, Action={action_str}, Timestamp={timestamp}")

        message = f"Hello {user_name}, your {action_str} was recorded at {timestamp}."
        send_sms(message)

        # Log to database as well
        log_to_db(user_id, user_name, action_str)

    except PermissionError:
        print(f"Permission denied. Please close the file '{EXCEL_FILE}' and try again.")
    except Exception as e:
        print(f"Error saving data to the file: {e}")

# Function to handle biometric events when a user is detected
def on_biometric_event(conn):
    print("Waiting for biometric event...")
    while True:
        try:
            attendance = conn.get_attendance()
            if attendance:
                for record in attendance:
                    user_id = record.user_id
                    action = record.status
                    if action in VALID_ACTIONS:
                        log_to_excel_and_send_sms(user_id, action)
            time.sleep(1)
        except Exception as e:
            print(f"Error reading attendance data: {e}")
            time.sleep(1)

# Main function to connect to the device and start monitoring
def main():
    try:
        print("Connecting to ZKTeco device...")
        zk = ZK(ZK_IP, port=ZK_PORT, timeout=5)
        conn = zk.connect()
        conn.disable_device()
        print("Connected successfully!")
        on_biometric_event(conn)
    except Exception as e:
        print(f"Error: {e}")
    finally:
        if conn:
            conn.disconnect()
        print("Disconnected from ZKTeco device.")

if __name__ == "__main__":
    main()
